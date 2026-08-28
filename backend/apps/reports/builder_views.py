"""Advanced Report Builder - Dynamic report creation with column selection, filters, grouping."""

from decimal import Decimal
from django.db.models import Count, Q, Case, When, Value, IntegerField, Sum, Avg, Max, Min
from django.db.models.fields import Field
from django.utils.module_loading import import_string
from django.utils import timezone
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from apps.accounts.access import apply_campus_scope, institution_scope
from apps.accounts.permissions import IsAccountantRole
from apps.reports.base_views import BaseReportView
from apps.reports.models import (
    CustomReportDataSource, CustomReport, SavedReport, ReportDefinition,
    ReportCategory
)
from apps.reports.utils import quantize, to_csv
from apps.reports.filters import (
    apply_filters, get_model_filter_specs, ReportFilterSpec,
    FilterOperator, FIELD_OPERATOR_MAP
)


class DataSourceListView(APIView):
    """List available data sources for report builder."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get(self, request):
        data_sources = CustomReportDataSource.objects.filter(
            is_active=True
        ).select_related("category").order_by("category__order", "name")

        user_roles = request.user.get_roles(institution=getattr(request, "institution", None))

        result = []
        for ds in data_sources:
            if ds.required_roles and not any(r in user_roles for r in ds.required_roles):
                continue
            result.append({
                "id": ds.id,
                "slug": ds.slug,
                "name": ds.name,
                "description": ds.description,
                "category": ds.category.name if ds.category else "-",
                "model_path": ds.model_path,
                "display_name_field": ds.display_name_field,
                "available_fields": ds.available_fields,
                "default_filters": ds.default_filters,
                "relationships": ds.relationships,
            })

        return Response(result)


class DataSourceDetailView(APIView):
    """Get detailed data source configuration."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get(self, request, pk):
        from apps.reports.models import CustomReportDataSource
        try:
            ds = CustomReportDataSource.objects.get(pk=pk, is_active=True)
        except CustomReportDataSource.DoesNotExist:
            return Response({"detail": "Data source not found"}, status=404)

        return Response({
            "id": ds.id,
            "slug": ds.slug,
            "name": ds.name,
            "description": ds.description,
            "category": ds.category.name if ds.category else "-",
            "model_path": ds.model_path,
            "display_name_field": ds.display_name_field,
            "id_field": ds.id_field,
            "available_fields": ds.available_fields,
            "default_filters": ds.default_filters,
            "relationships": ds.relationships,
            "select_related_fields": ds.select_related_fields,
            "prefetch_related_fields": ds.prefetch_related_fields,
        })


class ReportBuilderView(APIView):
    """Advanced report builder - create, preview, and execute custom reports."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get_base_queryset(self, data_source, request):
        """Get base queryset for a data source."""
        model = import_string(data_source.model_path)
        queryset = model.objects.all()

        if data_source.select_related_fields:
            queryset = queryset.select_related(*data_source.select_related_fields)

        if data_source.prefetch_related_fields:
            queryset = queryset.prefetch_related(*data_source.prefetch_related_fields)

        # Apply institution scope
        queryset = institution_scope(queryset, request, "institution_id")

        # Apply default filters from data source
        if data_source.default_filters:
            queryset = apply_filters(queryset, data_source.default_filters, model)

        return queryset

    def apply_joins(self, queryset, primary_model, joins, request):
        """Apply joins to the queryset."""
        for join in joins:
            related_ds = join.get("data_source")
            join_type = join.get("join_type", "LEFT")  # LEFT, INNER
            on_clause = join.get("on")  # {"left_field": "right_field"}

            if not related_ds or not on_clause:
                continue

            try:
                related_model = import_string(related_ds)
            except ImportError:
                continue

            # For now, we use select_related/prefetch_related for simple joins
            # Complex joins would require raw SQL or more sophisticated ORM
            for left_field, right_field in on_clause.items():
                if join_type == "INNER":
                    queryset = queryset.filter(**{f"{left_field}__isnull": False})

        return queryset

    def execute_report(self, report_config, request):
        """Execute a custom report and return results."""
        primary_ds_slug = report_config.get("primary_data_source")
        try:
            primary_ds = CustomReportDataSource.objects.get(slug=primary_ds_slug, is_active=True)
        except CustomReportDataSource.DoesNotExist:
            return {"error": "Primary data source not found"}

        model = import_string(primary_ds.model_path)

        # Build queryset
        queryset = self.get_base_queryset(primary_ds, request)

        # Apply joins
        joins = report_config.get("joined_data_sources", [])
        queryset = self.apply_joins(queryset, model, joins, request)

        # Apply filters
        filters = report_config.get("filters", {})
        filter_logic = report_config.get("filter_logic", "AND")
        if filters:
            queryset = apply_filters(queryset, filters, model)

        # Apply campus scope
        campus_field = "campus_id"
        for field in model._meta.get_fields():
            if hasattr(field, "related_model") and field.related_model and field.related_model.__name__ == "Campus":
                campus_field = field.name
                break
        queryset = apply_campus_scope(queryset, request, campus_field)

        # Apply sorting
        sorting = report_config.get("sorting", [])
        if sorting:
            order_by = []
            for sort in sorting:
                field = sort.get("field")
                direction = sort.get("direction", "asc")
                if field:
                    order_by.append(f"-{field}" if direction == "desc" else field)
            if order_by:
                queryset = queryset.order_by(*order_by)

        # Apply grouping/aggregation
        grouping = report_config.get("grouping", {})
        aggregations = report_config.get("aggregations", [])

        if grouping:
            group_by = grouping.get("fields", [])
            if group_by:
                agg_dict = {}
                for agg in aggregations:
                    agg_type = agg.get("type")  # count, sum, avg, max, min
                    field = agg.get("field")
                    alias = agg.get("alias", f"{field}__{agg_type}")

                    if agg_type == "count":
                        agg_dict[alias] = Count(field or "id")
                    elif agg_type == "sum" and field:
                        agg_dict[alias] = Sum(field)
                    elif agg_type == "avg" and field:
                        agg_dict[alias] = Avg(field)
                    elif agg_type == "max" and field:
                        agg_dict[alias] = Max(field)
                    elif agg_type == "min" and field:
                        agg_dict[alias] = Min(field)

                if agg_dict:
                    queryset = queryset.values(*group_by).annotate(**agg_dict)
                else:
                    queryset = queryset.values(*group_by).annotate(count=Count("id"))

        # Get selected fields
        selected_fields = report_config.get("selected_fields", [])
        field_paths = [f.get("path") for f in selected_fields if f.get("path")]

        # Apply pagination
        page = int(request.query_params.get("page", 1))
        page_size = min(int(request.query_params.get("page_size", 50)), 500)

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size

        # Get results
        if field_paths:
            results = list(queryset.values(*field_paths)[start:end])
        else:
            results = list(queryset[start:end])

        # Format results with custom labels
        formatted_results = []
        for row in results:
            formatted_row = {}
            for field_config in selected_fields:
                path = field_config.get("path")
                label = field_config.get("label", path)
                if path in row:
                    formatted_row[label] = row[path]
            formatted_results.append(formatted_row)

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "results": formatted_results,
            "fields": [f.get("label", f.get("path")) for f in selected_fields],
        }

    def post(self, request):
        """Preview or execute a custom report."""
        action = request.data.get("action", "preview")  # preview, execute, save
        report_config = request.data.get("report_config", {})

        if action == "save":
            return self.save_report(request, report_config)

        result = self.execute_report(report_config, request)

        if "error" in result:
            return Response(result, status=400)

        if action == "export":
            return self.export_report(result, request)

        return Response(result)

    def save_report(self, request, report_config):
        """Save a custom report configuration."""
        primary_ds_slug = report_config.get("primary_data_source")
        try:
            primary_ds = CustomReportDataSource.objects.get(slug=primary_ds_slug, is_active=True)
        except CustomReportDataSource.DoesNotExist:
            return Response({"detail": "Primary data source not found"}, status=400)

        report = CustomReport.objects.create(
            name=request.data.get("name", "Untitled Report"),
            description=request.data.get("description", ""),
            primary_data_source=primary_ds,
            joined_data_sources=report_config.get("joined_data_sources", []),
            selected_fields=report_config.get("selected_fields", []),
            filters=report_config.get("filters", {}),
            filter_logic=report_config.get("filter_logic", "AND"),
            grouping=report_config.get("grouping", {}),
            sorting=report_config.get("sorting", []),
            aggregations=report_config.get("aggregations", []),
            template_id=request.data.get("template_id"),
            created_by=request.user,
        )

        return Response({
            "id": report.id,
            "name": report.name,
            "detail": "Report saved successfully",
        }, status=201)

    def export_report(self, result, request):
        """Export report results as CSV/Excel."""
        fmt = request.query_params.get("format", "csv")
        filename = request.query_params.get("filename", "custom_report")

        if fmt == "csv":
            return self.export_csv(result, filename)
        elif fmt == "excel":
            return self.export_excel(result, filename)
        return Response({"detail": "Invalid format"}, status=400)

    def export_csv(self, result, filename):
        headers = result.get("fields", [])
        rows = []
        for row in result.get("results", []):
            rows.append([row.get(h, "") for h in headers])

        return to_csv(f"{filename}.csv", headers, rows)

    def export_excel(self, result, filename):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response({"detail": "openpyxl not installed"}, status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        headers = result.get("fields", [])
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(result.get("results", []), 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        from django.http import HttpResponse
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response


class ReportBuilderConfigView(APIView):
    """Get configuration for report builder UI."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get(self, request):
        # Get all categories with their data sources
        categories = ReportCategory.objects.filter(is_active=True).prefetch_related("data_sources")

        user_roles = request.user.get_roles(institution=getattr(request, "institution", None))

        result = []
        for cat in categories:
            sources = []
            for ds in cat.data_sources.filter(is_active=True):
                if ds.required_roles and not any(r in user_roles for r in ds.required_roles):
                    continue
                sources.append({
                    "id": ds.id,
                    "slug": ds.slug,
                    "name": ds.name,
                    "description": ds.description,
                    "model_path": ds.model_path,
                    "available_fields": ds.available_fields,
                    "default_filters": ds.default_filters,
                    "relationships": ds.relationships,
                })

            if sources:
                result.append({
                    "category": cat.name,
                    "slug": cat.slug,
                    "icon": cat.icon,
                    "data_sources": sources,
                })

        # Get available operators for each field type
        operators = {
            ft: ops for ft, ops in FIELD_OPERATOR_MAP.items()
        }

        return Response({
            "categories": result,
            "operators": operators,
            "aggregation_types": ["count", "sum", "avg", "max", "min"],
            "join_types": ["LEFT", "INNER"],
        })


class CustomReportListView(APIView):
    """List user's custom reports."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get(self, request):
        reports = CustomReport.objects.filter(
            created_by=request.user
        ).select_related("primary_data_source", "template").order_by("-is_favorite", "-updated_at")

        data = []
        for r in reports:
            data.append({
                "id": r.id,
                "name": r.name,
                "description": r.description,
                "primary_data_source": {
                    "slug": r.primary_data_source.slug,
                    "name": r.primary_data_source.name,
                },
                "is_favorite": r.is_favorite,
                "is_shared": r.is_shared,
                "created_at": r.created_at.isoformat(),
                "updated_at": r.updated_at.isoformat(),
                "last_run_at": r.last_run_at.isoformat() if r.last_run_at else None,
            })
        return Response(data)

    def post(self, request):
        return ReportBuilderView().save_report(request, request.data.get("report_config", {}))


class CustomReportDetailView(APIView):
    """Detail view for custom reports."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get_object(self, pk, user):
        try:
            return CustomReport.objects.get(pk=pk, created_by=user)
        except CustomReport.DoesNotExist:
            return None

    def get(self, request, pk):
        report = self.get_object(pk, request.user)
        if not report:
            return Response({"detail": "Not found"}, status=404)

        return Response({
            "id": report.id,
            "name": report.name,
            "description": report.description,
            "primary_data_source": {
                "id": report.primary_data_source.id,
                "slug": report.primary_data_source.slug,
                "name": report.primary_data_source.name,
            },
            "joined_data_sources": report.joined_data_sources,
            "selected_fields": report.selected_fields,
            "filters": report.filters,
            "filter_logic": report.filter_logic,
            "grouping": report.grouping,
            "sorting": report.sorting,
            "aggregations": report.aggregations,
            "template_id": report.template_id,
            "is_favorite": report.is_favorite,
            "is_shared": report.is_shared,
            "shared_with": list(report.shared_with.values_list("id", flat=True)),
            "created_at": report.created_at.isoformat(),
            "updated_at": report.updated_at.isoformat(),
            "last_run_at": report.last_run_at.isoformat() if report.last_run_at else None,
        })

    def put(self, request, pk):
        report = self.get_object(pk, request.user)
        if not report:
            return Response({"detail": "Not found"}, status=404)

        report.name = request.data.get("name", report.name)
        report.description = request.data.get("description", report.description)
        report.joined_data_sources = request.data.get("joined_data_sources", report.joined_data_sources)
        report.selected_fields = request.data.get("selected_fields", report.selected_fields)
        report.filters = request.data.get("filters", report.filters)
        report.filter_logic = request.data.get("filter_logic", report.filter_logic)
        report.grouping = request.data.get("grouping", report.grouping)
        report.sorting = request.data.get("sorting", report.sorting)
        report.aggregations = request.data.get("aggregations", report.aggregations)
        report.template_id = request.data.get("template_id", report.template_id)
        report.is_favorite = request.data.get("is_favorite", report.is_favorite)
        report.is_shared = request.data.get("is_shared", report.is_shared)
        report.updated_by = request.user
        report.save()

        if "shared_with" in request.data:
            report.shared_with.set(request.data["shared_with"])

        return Response({"detail": "Report updated"})

    def delete(self, request, pk):
        report = self.get_object(pk, request.user)
        if not report:
            return Response({"detail": "Not found"}, status=404)
        report.delete()
        return Response({"detail": "Report deleted"})


class CustomReportExecuteView(APIView):
    """Execute a saved custom report."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def post(self, request, pk):
        from apps.reports.models import CustomReport
        try:
            report = CustomReport.objects.get(pk=pk, created_by=request.user)
        except CustomReport.DoesNotExist:
            return Response({"detail": "Not found"}, status=404)

        report_config = {
            "primary_data_source": report.primary_data_source.slug,
            "joined_data_sources": report.joined_data_sources,
            "selected_fields": report.selected_fields,
            "filters": report.filters,
            "filter_logic": report.filter_logic,
            "grouping": report.grouping,
            "sorting": report.sorting,
            "aggregations": report.aggregations,
        }

        builder = ReportBuilderView()
        result = builder.execute_report(report_config, request)

        if "error" in result:
            return Response(result, status=400)

        # Update last run
        report.last_run_at = timezone.now()
        report.run_count += 1
        report.save(update_fields=["last_run_at", "run_count"])

        return Response(result)