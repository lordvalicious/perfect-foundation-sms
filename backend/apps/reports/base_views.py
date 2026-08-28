"""Base classes and utilities for report generation."""

from abc import ABC, abstractmethod
from decimal import Decimal
from typing import Any, Dict, List, Optional

from django.db.models import QuerySet, Q, Count, Sum, Avg, Max, Min
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.access import apply_campus_scope, campus_access
from apps.accounts.permissions import IsAccountantRole
from apps.reports.filters import apply_filters, get_model_filter_specs, ReportFilterSpec


class ReportExportMixin:
    """Mixin to add CSV/Excel/PDF export capabilities to report views."""

    export_filename = "report"
    export_headers = []
    export_fields = []

    def get_export_queryset(self, request):
        """Override to provide the queryset for export."""
        return self.get_queryset(request)

    def get_export_data(self, request):
        """Override to provide custom export data."""
        queryset = self.get_export_queryset(request)
        data = []
        for obj in queryset:
            row = []
            for field in self.export_fields:
                value = self.get_field_value(obj, field)
                row.append(value)
            data.append(row)
        return data

    def get_field_value(self, obj, field_path: str):
        """Get a field value from an object using dot notation."""
        current = obj
        for part in field_path.split("__"):
            if current is None:
                return ""
            current = getattr(current, part, None)
            if callable(current):
                current = current()
        return current or ""

    def export_csv(self, request):
        """Export report as CSV."""
        import csv

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{self.export_filename}.csv"'
        response.write("\ufeff")

        writer = csv.writer(response)
        writer.writerow(self.export_headers)
        for row in self.get_export_data(request):
            writer.writerow(row)

        return response

    def export_excel(self, request):
        """Export report as Excel (xlsx)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("openpyxl not installed", status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(self.export_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(self.get_export_data(request), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx in range(1, len(self.export_headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.export_filename}.xlsx"'
        wb.save(response)
        return response

    def finalize_response(self, request, response, *args, **kwargs):
        fmt = request.query_params.get("format")
        if fmt == "csv":
            return self.export_csv(request)
        elif fmt == "excel":
            return self.export_excel(request)
        return super().finalize_response(request, response, *args, **kwargs)


class BaseReportView(APIView, ReportExportMixin):
    """
    Base class for all report views.

    Subclasses should define:
    - report_definition_key: The key of the ReportDefinition
    - get_queryset(request): Return the base queryset
    - get_serializer_context(request): Optional context for serialization
    """

    permission_classes = [IsAuthenticated, IsAccountantRole]
    report_definition_key = None
    model = None
    default_filters = {}
    available_filters = []
    default_columns = []
    available_columns = []
    grouping_fields = []
    supports_pagination = True
    page_size = 50
    max_page_size = 500

    def get_report_definition(self):
        from apps.reports.models import ReportDefinition
        if self.report_definition_key:
            try:
                return ReportDefinition.objects.get(key=self.report_definition_key, is_active=True)
            except ReportDefinition.DoesNotExist:
                pass
        return None

    def get_base_queryset(self, request):
        """Get the base queryset before filters. Override in subclasses."""
        if self.model:
            return self.model.objects.all()
        return QuerySet(model=self.model)

    def get_queryset(self, request):
        """Get the filtered, scoped queryset."""
        queryset = self.get_base_queryset(request)

        # Apply institution and campus scoping
        if self.model:
            queryset = apply_campus_scope(queryset, request, "campus_id")

        # Apply filters from query params
        filters = self.parse_request_filters(request)
        if filters:
            queryset = apply_filters(queryset, filters, self.model)

        # Apply default filters
        for key, value in self.default_filters.items():
            if key not in filters:
                queryset = queryset.filter(**{key: value})

        return queryset

    def parse_request_filters(self, request) -> dict:
        """Parse filters from request query params."""
        filters = {}
        for param, value in request.query_params.items():
            if param in ("format", "page", "page_size", "sort", "group_by", "columns"):
                continue
            if value:
                filters[param] = {"operator": "exact", "value": value}
        return filters

    def apply_sorting(self, queryset, request):
        """Apply sorting from request."""
        sort = request.query_params.get("sort", "")
        if sort:
            sort_fields = []
            for field in sort.split(","):
                if field.startswith("-"):
                    sort_fields.append(field)
                else:
                    sort_fields.append(field)
            if sort_fields:
                queryset = queryset.order_by(*sort_fields)
        return queryset

    def apply_grouping(self, queryset, request):
        """Apply grouping/aggregation."""
        group_by = request.query_params.get("group_by", "")
        if group_by and group_by in self.grouping_fields:
            group_field = group_by
            aggregations = request.query_params.get("aggregations", "count")
            agg_funcs = {}
            for agg in aggregations.split(","):
                if agg == "count":
                    agg_funcs["count"] = Count("id")
                elif agg.startswith("sum:"):
                    agg_funcs[agg] = Sum(agg.split(":")[1])
                elif agg.startswith("avg:"):
                    agg_funcs[agg] = Avg(agg.split(":")[1])
                elif agg.startswith("max:"):
                    agg_funcs[agg] = Max(agg.split(":")[1])
                elif agg.startswith("min:"):
                    agg_funcs[agg] = Min(agg.split(":")[1])

            queryset = queryset.values(group_field).annotate(**agg_funcs).order_by(group_field)
        return queryset

    def get_paginated_response(self, queryset, request):
        """Return paginated response."""
        if not self.supports_pagination:
            return queryset

        page = int(request.query_params.get("page", 1))
        page_size = min(
            int(request.query_params.get("page_size", self.page_size)),
            self.max_page_size
        )

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size

        return {
            "count": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "results": queryset[start:end],
        }

    def get(self, request):
        """Handle GET request - return report data."""
        queryset = self.get_queryset(request)
        queryset = self.apply_sorting(queryset, request)

        if request.query_params.get("group_by"):
            queryset = self.apply_grouping(queryset, request)
            return Response({"results": list(queryset)})

        paginated = self.get_paginated_response(queryset, request)
        if isinstance(paginated, dict):
            return Response(paginated)

        return Response({"results": list(paginated)})


class AggregateReportView(BaseReportView):
    """Base class for aggregated reports (summary + details)."""

    def get_summary(self, queryset, request) -> Dict[str, Any]:
        """Calculate summary statistics. Override in subclasses."""
        return {"total": queryset.count()}

    def get_detail_rows(self, queryset, request) -> List[Dict[str, Any]]:
        """Get detail rows. Override in subclasses."""
        return []

    def get(self, request):
        queryset = self.get_queryset(request)
        summary = self.get_summary(queryset, request)
        rows = self.get_detail_rows(queryset, request)

        return Response({
            "summary": summary,
            "results": rows,
        })


class ReportConfigView(APIView):
    """API endpoint to get report configuration (filters, columns, etc.)."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get(self, request, report_key):
        from apps.reports.models import ReportDefinition

        try:
            report = ReportDefinition.objects.get(key=report_key, is_active=True)
        except ReportDefinition.DoesNotExist:
            return Response({"detail": "Report not found"}, status=404)

        return Response({
            "key": report.key,
            "title": report.title,
            "description": report.description,
            "endpoint_url": report.endpoint_url,
            "supports_csv": report.supports_csv,
            "supports_pdf": report.supports_pdf,
            "supports_excel": report.supports_excel,
            "supports_print": report.supports_print,
            "supports_schedule": report.supports_schedule,
            "default_filters": report.default_filters,
            "available_filters": report.available_filters,
            "default_columns": report.default_columns,
            "available_columns": report.available_columns,
            "supports_grouping": report.supports_grouping,
            "grouping_fields": report.grouping_fields,
            "supports_drilldown": report.supports_drilldown,
            "drilldown_target": report.drilldown_target,
        })


class ReportListView(APIView):
    """List all available reports grouped by category."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.reports.models import ReportCategory, ReportDefinition

        categories = ReportCategory.objects.filter(is_active=True).prefetch_related("report_definitions")

        user = request.user
        user_roles = user.get_roles(institution=getattr(request, "institution", None))

        result = []
        for category in categories:
            if category.required_roles and not any(r in user_roles for r in category.required_roles):
                continue

            reports = []
            for report in category.report_definitions.filter(is_active=True):
                if report.required_roles and not any(r in user_roles for r in report.required_roles):
                    continue
                reports.append({
                    "key": report.key,
                    "title": report.title,
                    "description": report.description,
                    "report_type": report.report_type,
                    "endpoint_url": report.endpoint_url,
                    "supports_csv": report.supports_csv,
                    "supports_pdf": report.supports_pdf,
                    "supports_excel": report.supports_excel,
                    "supports_print": report.supports_print,
                    "supports_schedule": report.supports_schedule,
                    "is_featured": report.is_featured,
                })

            if reports:
                result.append({
                    "category": category.name,
                    "slug": category.slug,
                    "icon": category.icon,
                    "description": category.description,
                    "reports": reports,
                })

        return Response(result)


class SavedReportView(APIView):
    """CRUD for saved reports."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get(self, request):
        from apps.reports.models import SavedReport

        reports = SavedReport.objects.filter(
            created_by=request.user
        ).select_related("report_definition", "template").order_by("-is_favorite", "-updated_at")

        data = []
        for r in reports:
            data.append({
                "id": r.id,
                "name": r.name,
                "description": r.description,
                "report_definition": {
                    "key": r.report_definition.key,
                    "title": r.report_definition.title,
                },
                "filters": r.filters,
                "columns": r.columns,
                "grouping": r.grouping,
                "sorting": r.sorting,
                "is_favorite": r.is_favorite,
                "is_shared": r.is_shared,
                "created_at": r.created_at.isoformat(),
                "updated_at": r.updated_at.isoformat(),
                "last_run_at": r.last_run_at.isoformat() if r.last_run_at else None,
                "run_count": r.run_count,
            })
        return Response(data)

    def post(self, request):
        from apps.reports.models import SavedReport, ReportDefinition

        report_def_key = request.data.get("report_definition")
        try:
            report_def = ReportDefinition.objects.get(key=report_def_key, is_active=True)
        except ReportDefinition.DoesNotExist:
            return Response({"detail": "Invalid report definition"}, status=400)

        saved = SavedReport.objects.create(
            name=request.data.get("name", "Untitled Report"),
            description=request.data.get("description", ""),
            report_definition=report_def,
            filters=request.data.get("filters", {}),
            columns=request.data.get("columns", []),
            column_order=request.data.get("column_order", []),
            column_labels=request.data.get("column_labels", {}),
            grouping=request.data.get("grouping", {}),
            sorting=request.data.get("sorting", []),
            template_id=request.data.get("template_id"),
            created_by=request.user,
        )

        return Response({
            "id": saved.id,
            "name": saved.name,
            "detail": "Report saved successfully",
        }, status=201)


class SavedReportDetailView(APIView):
    """Detail view for saved reports."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get_object(self, pk, user):
        from apps.reports.models import SavedReport
        try:
            return SavedReport.objects.get(pk=pk, created_by=user)
        except SavedReport.DoesNotExist:
            return None

    def get(self, request, pk):
        saved = self.get_object(pk, request.user)
        if not saved:
            return Response({"detail": "Not found"}, status=404)

        return Response({
            "id": saved.id,
            "name": saved.name,
            "description": saved.description,
            "report_definition": {
                "key": saved.report_definition.key,
                "title": saved.report_definition.title,
            },
            "filters": saved.filters,
            "columns": saved.columns,
            "column_order": saved.column_order,
            "column_labels": saved.column_labels,
            "grouping": saved.grouping,
            "sorting": saved.sorting,
            "template_id": saved.template_id,
            "is_favorite": saved.is_favorite,
            "is_shared": saved.is_shared,
            "shared_with": list(saved.shared_with.values_list("id", flat=True)),
            "shared_with_roles": saved.shared_with_roles,
            "created_at": saved.created_at.isoformat(),
            "updated_at": saved.updated_at.isoformat(),
            "last_run_at": saved.last_run_at.isoformat() if saved.last_run_at else None,
            "run_count": saved.run_count,
        })

    def put(self, request, pk):
        saved = self.get_object(pk, request.user)
        if not saved:
            return Response({"detail": "Not found"}, status=404)

        saved.name = request.data.get("name", saved.name)
        saved.description = request.data.get("description", saved.description)
        saved.filters = request.data.get("filters", saved.filters)
        saved.columns = request.data.get("columns", saved.columns)
        saved.column_order = request.data.get("column_order", saved.column_order)
        saved.column_labels = request.data.get("column_labels", saved.column_labels)
        saved.grouping = request.data.get("grouping", saved.grouping)
        saved.sorting = request.data.get("sorting", saved.sorting)
        saved.template_id = request.data.get("template_id", saved.template_id)
        saved.is_favorite = request.data.get("is_favorite", saved.is_favorite)
        saved.is_shared = request.data.get("is_shared", saved.is_shared)
        saved.shared_with_roles = request.data.get("shared_with_roles", saved.shared_with_roles)
        saved.updated_by = request.user
        saved.save()

        if "shared_with" in request.data:
            saved.shared_with.set(request.data["shared_with"])

        return Response({"detail": "Report updated"})

    def delete(self, request, pk):
        saved = self.get_object(pk, request.user)
        if not saved:
            return Response({"detail": "Not found"}, status=404)
        saved.delete()
        return Response({"detail": "Report deleted"})


class ReportAuditView(APIView):
    """View report audit logs."""

    permission_classes = [IsAuthenticated, IsAccountantRole]

    def get(self, request):
        from apps.reports.models import ReportAuditLog

        logs = ReportAuditLog.objects.filter(user=request.user).order_by("-created_at")

        # Apply filters
        report_key = request.query_params.get("report")
        if report_key:
            logs = logs.filter(report_definition__key=report_key)

        action = request.query_params.get("action")
        if action:
            logs = logs.filter(action=action)

        date_from = request.query_params.get("date_from")
        if date_from:
            logs = logs.filter(created_at__date__gte=date_from)

        date_to = request.query_params.get("date_to")
        if date_to:
            logs = logs.filter(created_at__date__lte=date_to)

        page = int(request.query_params.get("page", 1))
        page_size = min(int(request.query_params.get("page_size", 50)), 200)

        total = logs.count()
        start = (page - 1) * page_size
        end = start + page_size

        data = []
        for log in logs[start:end]:
            data.append({
                "id": log.id,
                "report": log.report_definition.title if log.report_definition else "Custom",
                "action": log.action,
                "campus": log.campus_name,
                "academic_year": log.academic_year_name,
                "filters": log.filters_used,
                "record_count": log.record_count,
                "file_size": log.file_size,
                "created_at": log.created_at.isoformat(),
            })

        return Response({
            "count": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "results": data,
        })